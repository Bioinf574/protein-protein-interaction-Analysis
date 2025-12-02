import os
import pandas as pd
import networkx as nx
import plotly.graph_objects as go
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# === USER INPUT ===
input_file = input("Enter the path to your PPI file (.tsv or .csv): ").strip()
output_dir = "output"
os.makedirs(output_dir, exist_ok=True)

# === LOAD DATA ===
file_ext = os.path.splitext(input_file)[1].lower()
if file_ext == ".tsv":
    df = pd.read_csv(input_file, sep="\t")
elif file_ext == ".csv":
    df = pd.read_csv(input_file)
else:
    raise ValueError("Unsupported file type. Please use a .tsv or .csv file.")

print(f"[INFO] Loaded {len(df)} interactions from {file_ext.upper()} file.")
print(f"[INFO] Columns detected: {df.columns.tolist()}")

# === DETECT PROTEIN AND SCORE COLUMNS ===
interaction_cols = [col for col in df.columns if 'protein' in col.lower() or 'node' in col.lower()]
score_cols = [col for col in df.columns if 'score' in col.lower()]

if len(interaction_cols) >= 2:
    col_A, col_B = interaction_cols[:2]
else:
    raise ValueError("❌ Could not detect protein columns. File must have columns like 'protein1'/'protein2' or '#node1'/'node2'.")

score_col = score_cols[0] if score_cols else None

# === SMART FILTERING (auto-adjust thresholds) ===
if score_col:
    original_len = len(df)
    thresholds = [700, 400, 0]
    for t in thresholds:
        filtered = df[df[score_col] > t]
        if not filtered.empty:
            df = filtered
            print(f"[INFO] Using threshold {t} for {score_col} → {len(df)} interactions (from {original_len})")
            break
    else:
        print("⚠️ No valid interactions found even at lowest threshold.")
else:
    print("⚠️ No score column found; using all interactions.")

# === HANDLE EMPTY DATAFRAME ===
if df.empty:
    print("⚠️ No interactions passed the filter — cannot build a network.")
    empty_df = pd.DataFrame([["No interactions found after filtering"]], columns=["Message"])
    empty_path = os.path.join(output_dir, "network_summary.xlsx")
    empty_df.to_excel(empty_path, index=False)
    print(f"[OK] Empty summary saved to {empty_path}")
    exit(0)

# === BUILD NETWORK ===
if score_col:
    G = nx.from_pandas_edgelist(df, col_A, col_B, [score_col])
else:
    G = nx.from_pandas_edgelist(df, col_A, col_B)

print(f"[INFO] Graph created with {G.number_of_nodes()} nodes and {G.number_of_edges()} edges.")

# === COMPUTE NETWORK METRICS ===
degree_dict = dict(G.degree())
betweenness_dict = nx.betweenness_centrality(G)
clustering_dict = nx.clustering(G)

# === BUILD SUMMARY DATAFRAME ===
summary_df = pd.DataFrame({
    "Protein": list(G.nodes()),
    "Degree": [degree_dict[n] for n in G.nodes()],
    "Betweenness": [betweenness_dict[n] for n in G.nodes()],
    "Clustering": [clustering_dict[n] for n in G.nodes()],
    "Number_of_Interactions": [len(list(G.neighbors(n))) for n in G.nodes()]
})

# === SORT BY DEGREE & INTERACTION ===
summary_df = summary_df.sort_values(by=["Degree", "Number_of_Interactions"], ascending=[False, False])

# === CREATE HUB SHEETS ===
top5 = summary_df.nlargest(5, "Degree")
top10 = summary_df.nlargest(10, "Degree")

# === PROTEIN INTERACTIONS LIST ===
interacting_nodes = {node: list(G.neighbors(node)) for node in G.nodes()}
interaction_df = pd.DataFrame({
    "Protein": list(interacting_nodes.keys()),
    "Number_of_Interactions": [len(v) for v in interacting_nodes.values()],
    "Interacting_Proteins": [", ".join(v) for v in interacting_nodes.values()]
}).sort_values(by="Number_of_Interactions", ascending=False)

# === SAVE ALL SHEETS TO EXCEL ===
summary_path = os.path.join(output_dir, "network_summary.xlsx")
with pd.ExcelWriter(summary_path, engine="openpyxl") as writer:
    summary_df.to_excel(writer, sheet_name="Summary", index=False)
    top5.to_excel(writer, sheet_name="Top5_Hubs", index=False)
    top10.to_excel(writer, sheet_name="Top10_Hubs", index=False)
    interaction_df.to_excel(writer, sheet_name="Protein_Interaction", index=False)

print(f"[OK] Summary metrics saved to {summary_path}")

# === FORMAT EXCEL OUTPUT ===
wb = load_workbook(summary_path)
thin = Side(border_style="thin", color="000000")
border = Border(left=thin, right=thin, top=thin, bottom=thin)
header_fill = PatternFill(start_color="87CEEB", end_color="87CEEB", fill_type="solid")
header_font = Font(color="000000", bold=True, size=14)
center_alignment = Alignment(horizontal="center", vertical="center")

for ws in wb.worksheets:
    # Auto-fit columns
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        ws.column_dimensions[col_letter].width = max_length + 2

    # Format header row
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_alignment
        cell.border = border

    # Format all data cells (center align except Interacting_Proteins)
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            if ws.title != "Protein_Interaction" or cell.column_letter != 'C':
                cell.alignment = center_alignment
            cell.border = border

wb.save(summary_path)
print("[OK] Excel formatting applied successfully.")

# === VISUALIZATION (Plotly) ===
pos = nx.spring_layout(G, seed=42)

# Edges
edge_x, edge_y = [], []
for edge in G.edges():
    x0, y0 = pos[edge[0]]
    x1, y1 = pos[edge[1]]
    edge_x.extend([x0, x1, None])
    edge_y.extend([y0, y1, None])
edge_trace = go.Scatter(x=edge_x, y=edge_y, line=dict(width=0.5, color="#888"), hoverinfo='none', mode='lines')

# --- Node Hover Text (Fixed Formatting) ---
node_x, node_y, node_text = [], [], []
for node in G.nodes():
    x, y = pos[node]
    node_x.append(x)
    node_y.append(y)
    neighbors = interacting_nodes[node]
    grouped_neighbors = [", ".join(neighbors[i:i+5]) for i in range(0, len(neighbors), 5)]
    neighbor_str = "<br>".join(grouped_neighbors)
    hover_text = f"<b>{node}</b><br><b>Interacts with {len(neighbors)} proteins:</b><br>{neighbor_str}"
    node_text.append(hover_text)

# --- Betweenness Plot ---
node_colors = [betweenness_dict[n] for n in G.nodes()]
node_trace = go.Scatter(
    x=node_x, y=node_y, mode='markers+text', text=list(G.nodes()),
    hovertext=node_text, hoverinfo="text", textposition="top center",
    marker=dict(
        showscale=True, colorscale='Viridis', color=node_colors,
        size=[8 + 2.5 * degree_dict[n] for n in G.nodes()],
        colorbar=dict(title="Betweenness Centrality"), line_width=1
    )
)
fig1 = go.Figure(data=[edge_trace, node_trace])
fig1.update_layout(title="Protein Interaction Network (Betweenness Centrality)",
                   showlegend=False, hovermode='closest',
                   margin=dict(b=0, l=0, r=0, t=40))
fig1.write_html(os.path.join(output_dir, "protein_network_betweenness.html"))

# --- Clustering Coefficient Plot ---
node_colors2 = [clustering_dict[n] for n in G.nodes()]
node_trace2 = go.Scatter(
    x=node_x, y=node_y, mode='markers+text', text=list(G.nodes()),
    hovertext=node_text, hoverinfo="text", textposition="top center",
    marker=dict(
        showscale=True, colorscale='Viridis', color=node_colors2,
        size=[8 + 2.5 * degree_dict[n] for n in G.nodes()],
        colorbar=dict(title="Clustering Coefficient"), line_width=1
    )
)
fig2 = go.Figure(data=[edge_trace, node_trace2])
fig2.update_layout(title="Protein Interaction Network (Clustering Coefficient)",
                   showlegend=False, hovermode='closest',
                   margin=dict(b=0, l=0, r=0, t=40))
fig2.write_html(os.path.join(output_dir, "protein_network_clustering.html"))

# === LINE PLOT FOR NUMBER OF INTERACTIONS ===
line_fig = go.Figure()
line_fig.add_trace(go.Scatter(
    x=interaction_df["Protein"], y=interaction_df["Number_of_Interactions"],
    mode="lines+markers", marker=dict(color="royalblue"),
    text=interaction_df["Interacting_Proteins"],
    hovertemplate="<b>%{x}</b><br>Interactions: %{y}<extra></extra>"
))
line_fig.update_layout(
    title="Protein vs Number of Interactions (Line Plot)",
    xaxis_title="Protein", yaxis_title="Number of Interactions",
    template="plotly_white"
)
line_fig.write_html(os.path.join(output_dir, "interaction_line_plot.html"))

print("\n✅ All analysis complete!")
print(f"📊 Output saved to: {summary_path}")
print(f"🌐 Network HTMLs and line plot saved in: {output_dir}")
